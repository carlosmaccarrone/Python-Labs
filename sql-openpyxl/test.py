import openpyxl
from datetime import date
import sqlite3


doc = openpyxl.load_workbook('prueba.xlsx')
pagina = doc.active
cantfilas = pagina.max_row

db = sqlite3.connect("deudores.db")
ptr = db.cursor()
SQL= "INSERT INTO deudores (ID, FECHA, DNI, CARTERA, PAGO, PLAN, CUOTA, MONTO, CANAL_PAGO, POR, HONORA, IVA, A_RENDIR, REND, FECHA2, OP, FAC, N_de_operacion, N_de_sucursal, SUCURSAL, IMP, OBSERVACION, BANCO, TRAMO) VALUES (NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);"



for i in xrange(2, cantfilas+1):
        data =  []
        for j in xrange(1, 24):
                #print(str(j)+str(type(pagina.cell(row = i, column = j).value)))
                if ( isinstance(pagina.cell(row = i, column = j).value , type(None)) ):
                    aux = None
                    print(str(j)+"NULO")
                else:
                    if ( isinstance(pagina.cell(row = i, column = j).value , date)):
                        aux = str(pagina.cell(row = i, column = j).value)
                        #print(str(pagina.cell(row = i, column = j).value))
                    elif( isinstance(pagina.cell(row = i, column = j).value, long)):
                        aux = float(pagina.cell(row = i, column = j).value)
                        #print(str(pagina.cell(row = i, column = j).value))
                    elif( isinstance(pagina.cell(row = i, column = j).value, float)):
                        aux = pagina.cell(row = i, column = j).value
                        #print(str(pagina.cell(row = i, column = j).value))
                    elif( isinstance(pagina.cell(row = i, column = j).value, str)):
                        aux = str(pagina.cell(row = i, column = j).value)
                        #print(str(pagina.cell(row = i, column = j).value))
                    elif( isinstance(pagina.cell(row = i, column = j).value, unicode)):
                        aux = str(pagina.cell(row = i, column = j).value)
                        #print(str(pagina.cell(row = i, column = j).value))
                data.append(aux)
        print(data)
        print(SQL,data)
        ptr.execute(SQL,data)

db.commit()
db.close()
