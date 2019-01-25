import openpyxl
import MySQLdb as mariadb
from datetime import date

doc = openpyxl.load_workbook('prueba.xlsx')
pagina = doc.active
cantfilas = pagina.max_row

db = mariadb.connect(host= "127.0.0.1",
                    user="cmaccarrone",
                    passwd="r00t",
                    db="prueba")
ptr = db.cursor()
SQL="INSERT INTO `prueba`.`deudores` "
SQL= SQL + "(`ID`, `FECHA`, `DNI`, `CARTERA`, `PAGO`, `PLAN`, `CUOTA`, `MONTO`, `CANAL_PAGO`, `POR`, `HONORA`, `IVA`, `A_RENDIR`, `REND`, `FECHA2`, `OP`, `FAC`, `N_de_operacion`, `N_de_sucursal`, `SUCURSAL`,`IMP`, `OBSERVACION`, `BANCO`, `TRAMO`) "
QUERY= SQL + "VALUES(NULL"

for i in xrange(2, cantfilas+1):
        data = ""
        for j in xrange(1, 24):
                if ( isinstance(pagina.cell(row = i, column = j).value , type(None)) ):
                    data = data + ", NULL"
                    #print("NULO")
                else:
                    if ( isinstance(pagina.cell(row = i, column = j).value , date)):
                        data = data + ",'"+ str(pagina.cell(row = i, column = j).value) + "'"
                        #print(str(pagina.cell(row = i, column = j).value))
                    elif( isinstance(pagina.cell(row = i, column = j).value, long)):
                        data = data + ","+str(pagina.cell(row = i, column = j).value)
                        #print(str(pagina.cell(row = i, column = j).value))
                    elif( isinstance(pagina.cell(row = i, column = j).value, float)):
                        data = data + ",'"+str(pagina.cell(row = i, column = j).value) + "'"
                        #print(str(pagina.cell(row = i, column = j).value))
                    elif( isinstance(pagina.cell(row = i, column = j).value, str)):
                        data = data + ",'"+ str(pagina.cell(row = i, column = j).value) + "'"
                        #print(str(pagina.cell(row = i, column = j).value))
                    elif( isinstance(pagina.cell(row = i, column = j).value, unicode)):
                        data = data + ",'"+ str(pagina.cell(row = i, column = j).value) + "'"
                        #print(str(pagina.cell(row = i, column = j).value))
        ptr.execute(QUERY+data+');')
        #print str(QUERY)+str(data)+');'
db.commit()
db.close()
